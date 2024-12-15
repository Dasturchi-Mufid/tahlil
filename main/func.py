from django.http.response import HttpResponse
import datetime,calendar
import openpyxl
from decimal import Decimal

def decimal_to_float(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError("Type not serializable")

def download_type_detail_xlsx(request,month,branch, data,tur):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # sheet.title = 'People Data'
    first_header = ['',month,branch]

    sheet.append(first_header)
    # Set the headers for the columns
    headers = ['№', 'Tur', 'Mahsulot', 'Miqdor','Narx','Summa','Kirim sana','Chiqim sana','O`rtacha kun']
    
    # Write the headers to the first row
    sheet.append(headers)

    # Ensure that data is not empty and properly structured
    if not data:
        print("No data to write to the file.")
    
    # Write data rows
    for i,row in enumerate(data,1):
        # Ensure the row is a dictionary and append values to the Excel sheet
        sheet.append([
            i,
            row.get('type', ''),
            row.get('product',''),  # Ensure phone has at least one element
            row.get('quantity',''),  # Ensure there's a second phone number
            row.get('price', ''),
            row.get('total', ''),
            row.get('income', ''),
            row.get('out', ''),
            row.get('difference', '')
        ])
    # Set the column widths (optional)
    column_widths = [5,15, 45, 10, 15, 15,15,15,15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Create a response object with 'Content-Type' for Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        charset='utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{branch}_{month}-{tur}.xlsx"'

    try:
        # Save the workbook to the response object
        workbook.save(response)
    except Exception as e:
        print(f"Error saving workbook: {e}")
        response = HttpResponse("Error generating file.", status=500)

    return response


def download_type_xlsx(request,month,branch, data):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # sheet.title = 'People Data'
    first_header = ['',month,branch]

    sheet.append(first_header)
    # Set the headers for the columns
    headers = ['ID', 'Tur', 'Miqdor', 'Summa','Foiz']
    
    # Write the headers to the first row
    sheet.append(headers)

    # Ensure that data is not empty and properly structured
    if not data:
        print("No data to write to the file.")
    
    # Write data rows
    for row in data:
        # Ensure the row is a dictionary and append values to the Excel sheet
        sheet.append([
            row.get('id', ''),
            row.get('type_name', ''),
            row.get('quantity',''),  # Ensure phone has at least one element
            row.get('sum',''),  # Ensure there's a second phone number
            row.get('percentage', '')
        ])
    # Set the column widths (optional)
    column_widths = [10, 45, 15, 15, 11]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Create a response object with 'Content-Type' for Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        charset='utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{branch}_{month}-turlar.xlsx"'

    try:
        # Save the workbook to the response object
        workbook.save(response)
    except Exception as e:
        print(f"Error saving workbook: {e}")
        response = HttpResponse("Error generating file.", status=500)

    return response


def get_items_by_id(data, target_id):
    result = []
    for month, items in data.items():
        # Filter items with the matching id
        filtered_items = [item for item in items if item['id'] == target_id]
        result.extend(filtered_items)  # Add the filtered items to the result list
    return result

# import openpyxl
# from django.http import HttpResponse

import openpyxl
from django.http import HttpResponse

import openpyxl
from django.http import HttpResponse

import openpyxl
from django.http import HttpResponse

def download_report_xlsx(request, month, branch, data, product_types):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # sheet.title = 'People Data'

    # Write the first row (branch and month)
    header_1 = ['', branch] + month
    sheet.append(header_1)

    # Construct the second row with merged cells for the headers
    header_2 = ['Tur']  # 'Tur' is in the first column
    # Add empty cells before each data key for proper alignment
    for m in data.keys():
        header_2 += ['', '', m]  # Add empty cells for each data key

    # Write the headers to the second row
    sheet.append(header_2)

    # Create the third row with 'Soni', 'Narx', '%'
    header_3 = [''] + ['Soni', 'Narx', '%'] * len(data.keys())
    sheet.append(header_3)

    # Write the data for each product type
    for t in product_types:
        row = [t['type_name']]  # Start with the product type name
        info = get_items_by_id(data=data, target_id=t['id'])
        for j in info:
            row += [j['quantity'], j['sum'], j['percentage']]  # Add the quantity, sum, and percentage
        if len(row) != len(header_3):
            dif = len(header_3) - len(row)
            row += [0] * dif  # Fill the row with zeros if the length doesn't match
        sheet.append(row)

    # Merge cells for each header (every 3 columns for each m)
    col_index = 2  # Start from the second column (after 'Tur')
    for m in data.keys():
        # Merge cells (each key should span 3 columns)
        sheet.merge_cells(start_row=2, start_column=col_index, end_row=2, end_column=col_index + 2)
        # Place the header value (e.g., 'Category1') in the first of the merged cells
        sheet.cell(row=2, column=col_index, value=m)
        col_index += 3  # Move to the next set of 3 columns

    # Set the column widths (optional)
    column_widths = [50] + [20] * len(header_3)
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Ensure that data is not empty
    if not data:
        print("No data to write to the file.")

    # Create a response object with 'Content-Type' for Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        charset='utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{branch}_{month}-turlar.xlsx"'

    try:
        # Save the workbook to the response object
        workbook.save(response)
    except Exception as e:
        print(f"Error saving workbook: {e}")
        response = HttpResponse("Error generating file.", status=500)

    return response



def get_date(month: datetime.date):
    # Ensure month is a datetime.date object
    if isinstance(month, str):
        # If month is a string, convert it to a datetime.date object
        month = datetime.datetime.strptime(month, "%Y-%m").date()

    # First day of the month
    first_day = month.replace(day=1)

    # Correct calculation for the next month
    # If month is December (12), the next month should be January of the next year
    if month.month == 12:
        next_month = datetime.date(month.year + 1, 1, 1)
    else:
        next_month = month.replace(month=month.month + 1, day=1)

    # Subtract one day from the first day of the next month to get the last day of the current month
    last_day = next_month - datetime.timedelta(days=1)

    return first_day, last_day

def get_data(query,db,params=[]):
    
    db.execute(query,params)
    data = db.fetchall()
    return data



def get_date_range(start,end):
    # Convert to datetime objects
    start_date = datetime.datetime.strptime(start, "%Y-%m")
    end_date = datetime.datetime.strptime(end, "%Y-%m")

    # Get the first day of the start month
    first_day = start_date.replace(day=1)

    # Get the last day of the end month
    last_day = datetime.datetime(end_date.year, end_date.month, calendar.monthrange(end_date.year, end_date.month)[1])

    return first_day,last_day

def get_month_ranges(start_date, end_date):
    # Initialize the dictionary to store results
    month_range_dict = {}

    # Loop through each month in the range
    current_date = start_date
    while current_date <= end_date:
        # Get the first day of the current month
        first_day = current_date.replace(day=1)
        
        # Get the last day of the current month
        last_day = datetime.datetime(current_date.year, current_date.month, calendar.monthrange(current_date.year, current_date.month)[1])
        
        # Get the full month name
        month_name = current_date.strftime('%B')  # '%B' gives the full month name
        
        # Store in the dictionary with the full month name as key and (first_day, last_day) as value
        month_range_dict[month_name] = (first_day, last_day)
        
        # Move to the next month
        if current_date.month == 12:
            current_date = datetime.datetime(current_date.year + 1, 1, 1)
        else:
            current_date = datetime.datetime(current_date.year, current_date.month + 1, 1)

    return month_range_dict



